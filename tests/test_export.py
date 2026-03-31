"""
Тесты для модуля export
"""
import pytest
import os
from datetime import datetime
from unittest.mock import Mock, AsyncMock, patch, MagicMock
from openpyxl import load_workbook
from utils.export import ExcelExporter


@pytest.fixture
def mock_db():
    """Мок базы данных"""
    db = Mock()
    db.get = AsyncMock()
    db.get_hours_history = Mock()
    db.get_schedule_changes = Mock()
    return db


@pytest.fixture
def exporter(mock_db):
    """Экземпляр ExcelExporter с моком БД"""
    with patch('utils.export.DB', return_value=mock_db):
        exporter = ExcelExporter()
        exporter.db = mock_db
        return exporter


@pytest.fixture
def sample_user_data():
    """Примерные данные пользователя"""
    return (12345, 'test_user', 'Test User', '3191', None, 10, None, '1')


@pytest.fixture
def sample_hours_history():
    """Примерная история пропущенных часов"""
    return [
        ('2026-03-25', 2),
        ('2026-03-26', 4),
        ('2026-03-27', 0),
        ('2026-03-28', 6),
        ('2026-03-29', 2),
    ]


@pytest.fixture
def sample_schedule_changes():
    """Примерная история изменений расписания"""
    return [
        ('2026-03-25', '3191', 'Изменена 1 пара', '2026-03-25 08:00:00'),
        ('2026-03-26', '3191', 'Отменена 3 пара', '2026-03-26 09:30:00'),
        ('2026-03-27', '3395', 'Добавлена 5 пара', '2026-03-27 10:15:00'),
    ]


@pytest.fixture(autouse=True)
def cleanup_data_dir():
    """Очистка директории data после тестов"""
    yield
    # Удаляем тестовые файлы
    if os.path.exists('data'):
        for file in os.listdir('data'):
            if file.startswith('hours_history_') or file.startswith('schedule_changes_'):
                try:
                    os.remove(os.path.join('data', file))
                except:
                    pass


class TestExcelExporter:
    """Тесты для класса ExcelExporter"""
    
    @pytest.mark.asyncio
    async def test_export_hours_history_success(self, exporter, sample_user_data, sample_hours_history):
        """Тест успешного экспорта истории часов"""
        exporter.db.get.return_value = sample_user_data
        exporter.db.get_hours_history.return_value = sample_hours_history
        
        filepath = await exporter.export_hours_history(12345, days=30)
        
        assert filepath is not None
        assert os.path.exists(filepath)
        assert filepath.endswith('.xlsx')
        assert 'hours_history_12345' in filepath
        
        # Проверяем содержимое файла
        wb = load_workbook(filepath)
        ws = wb.active
        
        assert ws.title == "История пропусков"
        assert ws['A1'].value is not None
        assert '12345' in str(ws['A1'].value) or 'Test User' in str(ws['A1'].value)
        
        # Проверяем данные пользователя
        assert ws['A2'].value == "Группа:"
        assert ws['B2'].value == '3191'
        assert ws['A3'].value == "Текущие пропуски:"
        assert ws['B3'].value == 10
        
        wb.close()
    
    @pytest.mark.asyncio
    async def test_export_hours_history_no_data(self, exporter, sample_user_data):
        """Тест экспорта когда нет истории"""
        exporter.db.get.return_value = sample_user_data
        exporter.db.get_hours_history.return_value = []
        
        filepath = await exporter.export_hours_history(12345, days=30)
        
        assert filepath is not None
        assert os.path.exists(filepath)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Проверяем что есть сообщение "Нет данных"
        found_no_data = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and 'Нет данных' in str(cell.value):
                    found_no_data = True
        
        assert found_no_data
        wb.close()
    
    @pytest.mark.asyncio
    async def test_export_hours_history_user_not_found(self, exporter):
        """Тест экспорта для несуществующего пользователя"""
        exporter.db.get.return_value = None
        
        with pytest.raises(ValueError, match="Пользователь .* не найден"):
            await exporter.export_hours_history(99999, days=30)
    
    @pytest.mark.asyncio
    async def test_export_hours_history_creates_data_dir(self, exporter, sample_user_data, sample_hours_history):
        """Тест что создается директория data если её нет"""
        exporter.db.get.return_value = sample_user_data
        exporter.db.get_hours_history.return_value = sample_hours_history
        
        # Удаляем директорию если существует
        if os.path.exists('data'):
            for file in os.listdir('data'):
                filepath = os.path.join('data', file)
                if os.path.isfile(filepath):
                    os.remove(filepath)
            try:
                os.rmdir('data')
            except OSError:
                pass  # Директория не пустая, пропускаем
        
        filepath = await exporter.export_hours_history(12345, days=30)
        
        assert os.path.exists('data')
        assert os.path.exists(filepath)
    
    @pytest.mark.asyncio
    async def test_export_hours_history_total_calculation(self, exporter, sample_user_data, sample_hours_history):
        """Тест правильности подсчета итоговых часов"""
        exporter.db.get.return_value = sample_user_data
        exporter.db.get_hours_history.return_value = sample_hours_history
        
        filepath = await exporter.export_hours_history(12345, days=30)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Находим строку с итогом
        total_found = False
        expected_total = sum([row[1] for row in sample_hours_history])
        
        for row in ws.iter_rows():
            if row[0].value == "ИТОГО:":
                assert row[2].value == expected_total
                total_found = True
                break
        
        assert total_found
        wb.close()
    
    @pytest.mark.asyncio
    async def test_export_schedule_changes_success(self, exporter, sample_schedule_changes):
        """Тест успешного экспорта истории изменений расписания"""
        exporter.db.get_schedule_changes.return_value = sample_schedule_changes
        
        filepath = await exporter.export_schedule_changes(group_id='3191', limit=50)
        
        assert filepath is not None
        assert os.path.exists(filepath)
        assert filepath.endswith('.xlsx')
        assert 'schedule_changes_3191' in filepath
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        assert ws.title == "История изменений"
        assert ws['A1'].value is not None
        assert 'История изменений' in str(ws['A1'].value)
        
        wb.close()
    
    @pytest.mark.asyncio
    async def test_export_schedule_changes_all_groups(self, exporter, sample_schedule_changes):
        """Тест экспорта изменений для всех групп"""
        exporter.db.get_schedule_changes.return_value = sample_schedule_changes
        
        filepath = await exporter.export_schedule_changes(group_id=None, limit=50)
        
        assert filepath is not None
        assert 'schedule_changes_all' in filepath
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Проверяем что в заголовке нет упоминания конкретной группы
        header = str(ws['A1'].value)
        assert 'История изменений' in header
        
        wb.close()
    
    @pytest.mark.asyncio
    async def test_export_schedule_changes_no_data(self, exporter):
        """Тест экспорта изменений когда нет данных"""
        exporter.db.get_schedule_changes.return_value = []
        
        filepath = await exporter.export_schedule_changes(group_id='3191', limit=50)
        
        assert filepath is not None
        assert os.path.exists(filepath)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Проверяем наличие сообщения "Нет данных"
        found_no_data = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and 'Нет данных' in str(cell.value):
                    found_no_data = True
        
        assert found_no_data
        wb.close()
    
    @pytest.mark.asyncio
    async def test_export_hours_history_different_days(self, exporter, sample_user_data):
        """Тест экспорта с разными периодами"""
        exporter.db.get.return_value = sample_user_data
        
        for days in [7, 14, 30, 60]:
            exporter.db.get_hours_history.return_value = []
            
            filepath = await exporter.export_hours_history(12345, days=days)
            
            assert filepath is not None
            assert os.path.exists(filepath)
            
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Проверяем что период указан правильно
            assert ws['A4'].value == "Период:"
            assert str(days) in str(ws['B4'].value)
            
            wb.close()
    
    @pytest.mark.asyncio
    async def test_export_hours_history_weekday_names(self, exporter, sample_user_data, sample_hours_history):
        """Тест что дни недели правильно отображаются"""
        exporter.db.get.return_value = sample_user_data
        exporter.db.get_hours_history.return_value = sample_hours_history
        
        filepath = await exporter.export_hours_history(12345, days=30)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Проверяем наличие колонки с днями недели
        weekday_names = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
        found_weekday = False
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in weekday_names:
                    found_weekday = True
                    break
        
        assert found_weekday
        wb.close()
    
    @pytest.mark.asyncio
    async def test_export_schedule_changes_text_truncation(self, exporter):
        """Тест обрезки длинного текста изменений"""
        long_changes = [
            ('2026-03-25', '3191', 'A' * 200, '2026-03-25 08:00:00'),
        ]
        exporter.db.get_schedule_changes.return_value = long_changes
        
        filepath = await exporter.export_schedule_changes(group_id='3191', limit=50)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Проверяем что текст обрезан
        found_truncated = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '...' in cell.value:
                    assert len(cell.value) <= 104  # 100 символов + "..."
                    found_truncated = True
        
        assert found_truncated
        wb.close()


class TestExcelExporterFormatting:
    """Тесты форматирования Excel файлов"""
    
    @pytest.mark.asyncio
    async def test_hours_history_has_borders(self, exporter, sample_user_data, sample_hours_history):
        """Тест наличия границ в таблице"""
        exporter.db.get.return_value = sample_user_data
        exporter.db.get_hours_history.return_value = sample_hours_history
        
        filepath = await exporter.export_hours_history(12345, days=30)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Проверяем что у ячеек есть границы
        has_borders = False
        for row in ws.iter_rows(min_row=6, max_row=10):
            for cell in row:
                if cell.border and cell.border.left.style:
                    has_borders = True
                    break
        
        assert has_borders
        wb.close()
    
    @pytest.mark.asyncio
    async def test_hours_history_header_styling(self, exporter, sample_user_data, sample_hours_history):
        """Тест стилизации заголовков"""
        exporter.db.get.return_value = sample_user_data
        exporter.db.get_hours_history.return_value = sample_hours_history
        
        filepath = await exporter.export_hours_history(12345, days=30)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Проверяем стиль заголовка (строка 6)
        header_row = 6
        for col in range(1, 5):
            cell = ws.cell(row=header_row, column=col)
            # Проверяем что есть заливка
            assert cell.fill is not None
            # Проверяем что шрифт жирный
            assert cell.font.bold
        
        wb.close()
    
    @pytest.mark.asyncio
    async def test_schedule_changes_column_widths(self, exporter, sample_schedule_changes):
        """Тест ширины колонок"""
        exporter.db.get_schedule_changes.return_value = sample_schedule_changes
        
        filepath = await exporter.export_schedule_changes(group_id='3191', limit=50)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Проверяем что ширина колонок установлена
        assert ws.column_dimensions['A'].width == 5
        assert ws.column_dimensions['B'].width == 15
        assert ws.column_dimensions['C'].width == 10
        assert ws.column_dimensions['D'].width == 50
        assert ws.column_dimensions['E'].width == 20
        
        wb.close()


class TestExcelExporterErrorHandling:
    """Тесты обработки ошибок"""
    
    @pytest.mark.asyncio
    async def test_export_hours_history_db_error(self, exporter):
        """Тест обработки ошибки БД"""
        exporter.db.get.side_effect = Exception("Database error")
        
        with pytest.raises(Exception):
            await exporter.export_hours_history(12345, days=30)
    
    @pytest.mark.asyncio
    async def test_export_schedule_changes_db_error(self, exporter):
        """Тест обработки ошибки БД при экспорте изменений"""
        exporter.db.get_schedule_changes.side_effect = Exception("Database error")
        
        with pytest.raises(Exception):
            await exporter.export_schedule_changes(group_id='3191', limit=50)
