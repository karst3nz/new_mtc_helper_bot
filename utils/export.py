"""
Модуль экспорта данных в Excel и PDF
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from utils.db import DB
from utils.log import create_logger

logger = create_logger(__name__)

# Регистрация шрифтов Liberation для поддержки кириллицы
try:
    pdfmetrics.registerFont(TTFont('Liberation', '/usr/share/fonts/liberation/LiberationSerif-Regular.ttf'))
    pdfmetrics.registerFont(TTFont('Liberation-Bold', '/usr/share/fonts/liberation/LiberationSerif-Bold.ttf'))
    FONT_NAME = 'Liberation'
    logger.info("Шрифт Liberation успешно зарегистрирован")
except Exception as e:
    logger.warning(f"Не удалось загрузить Liberation: {e}, пробуем DejaVu")
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/TTF/DejaVuSerif.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVu-Bold', '/usr/share/fonts/TTF/DejaVuSerif-Bold.ttf'))
        FONT_NAME = 'DejaVu'
        logger.info("Шрифт DejaVu успешно зарегистрирован")
    except Exception as e2:
        logger.error(f"Не удалось загрузить шрифты: {e2}")
        FONT_NAME = 'Helvetica'


class ExcelExporter:
    def __init__(self):
        self.db = DB()
    
    async def export_hours_history(self, user_id: int, days: int = 30) -> str:
        """
        Экспорт истории пропущенных часов в Excel
        Возвращает путь к созданному файлу
        """
        logger.info(f"Экспорт истории часов для пользователя {user_id}")
        
        try:
            # Получаем данные пользователя
            user_data = await self.db.get(user_id)
            if not user_data:
                raise ValueError(f"Пользователь {user_id} не найден")
            
            # Получаем историю
            history = self.db.get_hours_history(user_id, days)
            
            # Создаем книгу Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "История пропусков"
            
            # Стили
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Заголовок документа
            ws.merge_cells('A1:D1')
            ws['A1'] = f"История пропущенных часов - Пользователь {user_data[2] or user_id}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Информация о пользователе
            ws['A2'] = "Группа:"
            ws['B2'] = user_data[3]
            ws['A3'] = "Текущие пропуски:"
            ws['B3'] = user_data[5]
            ws['A4'] = "Период:"
            ws['B4'] = f"Последние {days} дней"
            
            # Заголовки таблицы
            headers = ['№', 'Дата', 'Пропущено часов', 'День недели']
            ws.append([])  # Пустая строка
            header_row = 6
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Данные
            if history:
                weekdays = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
                
                for idx, row in enumerate(history, 1):
                    date_str = row[0]
                    hours = row[1]
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                    weekday = weekdays[date_obj.weekday()]
                    
                    data_row = [idx, date_obj.strftime("%d.%m.%Y"), hours, weekday]
                    ws.append(data_row)
                    
                    # Применяем границы
                    for col_num in range(1, len(headers) + 1):
                        ws.cell(row=header_row + idx, column=col_num).border = border
                        ws.cell(row=header_row + idx, column=col_num).alignment = Alignment(horizontal="center")
                
                # Итоговая строка
                total_row = header_row + len(history) + 1
                ws.cell(row=total_row, column=1).value = "ИТОГО:"
                ws.cell(row=total_row, column=1).font = Font(bold=True)
                ws.cell(row=total_row, column=3).value = sum([row[1] for row in history])
                ws.cell(row=total_row, column=3).font = Font(bold=True)
                
                for col_num in range(1, len(headers) + 1):
                    ws.cell(row=total_row, column=col_num).border = border
            else:
                ws.append(['', 'Нет данных за указанный период', '', ''])
            
            # Автоширина колонок
            for col in range(1, len(headers) + 1):
                max_length = 0
                column = get_column_letter(col)
                for cell in ws[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            
            # Сохранение файла
            filename = f"hours_history_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join("data", filename)
            
            # Создаем директорию если не существует
            os.makedirs("data", exist_ok=True)
            
            wb.save(filepath)
            logger.info(f"Файл успешно создан: {filepath}")
            
            return filepath
        
        except Exception as e:
            logger.error(f"Ошибка экспорта для пользователя {user_id}: {e}")
            raise
    
    async def export_schedule_changes(self, group_id: str = None, limit: int = 50) -> str:
        """
        Экспорт истории изменений расписания в Excel
        """
        logger.info(f"Экспорт истории изменений расписания для группы {group_id or 'все'}")
        
        try:
            # Получаем историю изменений
            changes = self.db.get_schedule_changes(group_id, limit)
            
            # Создаем книгу Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "История изменений"
            
            # Стили
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Заголовок
            ws.merge_cells('A1:E1')
            ws['A1'] = f"История изменений расписания" + (f" - Группа {group_id}" if group_id else "")
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Заголовки таблицы
            headers = ['№', 'Дата расписания', 'Группа', 'Изменения', 'Время записи']
            ws.append([])
            header_row = 3
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Данные
            if changes:
                for idx, row in enumerate(changes, 1):
                    date_str = row[0]
                    group = row[1]
                    diff_text = row[2][:100] + "..." if len(row[2]) > 100 else row[2]  # Обрезаем длинный текст
                    created_at = row[3]
                    
                    data_row = [idx, date_str, group, diff_text, created_at]
                    ws.append(data_row)
                    
                    # Применяем границы
                    for col_num in range(1, len(headers) + 1):
                        ws.cell(row=header_row + idx, column=col_num).border = border
                        if col_num != 4:  # Не центрируем текст изменений
                            ws.cell(row=header_row + idx, column=col_num).alignment = Alignment(horizontal="center")
            else:
                ws.append(['', '', 'Нет данных', '', ''])
            
            # Автоширина колонок
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 50
            ws.column_dimensions['E'].width = 20
            
            # Сохранение файла
            filename = f"schedule_changes_{group_id or 'all'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join("data", filename)
            
            os.makedirs("data", exist_ok=True)
            wb.save(filepath)
            logger.info(f"Файл успешно создан: {filepath}")
            
            return filepath
        
        except Exception as e:
            logger.error(f"Ошибка экспорта истории изменений: {e}")
            raise
    
    async def export_all_user_data(self, user_id: int) -> str:
        """
        Экспорт ВСЕХ данных пользователя в PDF документ
        Включает: профиль, настройки, историю часов, уведомления
        Возвращает путь к созданному файлу
        """
        logger.info(f"Экспорт всех данных для пользователя {user_id}")
        
        try:
            # Получаем все данные пользователя
            user_data = self.db.get_user_dataclass(user_id)
            if not user_data:
                raise ValueError(f"Пользователь {user_id} не найден")
            
            # Получаем дополнительные данные
            hours_history = self.db.get_hours_history(user_id, days=365)  # За год
            notification_settings = self.db.get_notification_settings(user_id)
            
            # Создаем PDF
            filename = f"user_data_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = os.path.join("data", filename)
            os.makedirs("data", exist_ok=True)
            
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            story = []
            
            # Стили - переопределяем с нашим шрифтом
            styles = getSampleStyleSheet()
            
            from reportlab.lib.styles import ParagraphStyle
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontName=FONT_NAME,
                fontSize=18,
                leading=22
            )
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontName=FONT_NAME,
                fontSize=14,
                leading=18
            )
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName=FONT_NAME,
                fontSize=10,
                leading=14
            )
            
            # Заголовок документа
            story.append(Paragraph("Экспорт данных пользователя", title_style))
            story.append(Paragraph(f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}", normal_style))
            story.append(Spacer(1, 0.5*cm))
            
            # 1. Основная информация
            story.append(Paragraph("1. Основная информация", heading_style))
            profile_data = [
                ['Параметр', 'Значение'],
                ['User ID', str(user_data.user_id)],
                ['Username', user_data.tg_username or 'Не указан'],
                ['Группа', str(user_data.group_id) if user_data.group_id else 'Не указана'],
                ['Доп. группа', str(user_data.sec_group_id) if user_data.sec_group_id else 'Не указана'],
                ['Смена', str(user_data.smena) if user_data.smena else 'Не указана'],
            ]
            
            profile_table = Table(profile_data, colWidths=[8*cm, 10*cm])
            profile_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(profile_table)
            story.append(Spacer(1, 0.5*cm))
            
            # 2. Пропущенные часы
            story.append(Paragraph("2. Пропущенные часы", heading_style))
            hours_data = [
                ['Параметр', 'Значение'],
                ['Текущие пропуски', str(user_data.missed_hours or 0)],
                ['Режим отображения', user_data.show_missed_hours_mode or 'Не настроен'],
            ]
            
            hours_table = Table(hours_data, colWidths=[8*cm, 10*cm])
            hours_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(hours_table)
            story.append(Spacer(1, 0.5*cm))
            
            # 3. История изменений
            if hours_history:
                story.append(Paragraph(f"3. История ({len(hours_history)} записей)", heading_style))
                
                history_data = [['№', 'Дата', 'Часы']]
                for idx, record in enumerate(hours_history[:50], 1):
                    date_val = record[0]
                    hours_val = record[1]
                    history_data.append([str(idx), date_val, str(hours_val)])
                
                history_table = Table(history_data, colWidths=[2*cm, 6*cm, 6*cm])
                history_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(history_table)
                story.append(Spacer(1, 0.5*cm))
            else:
                story.append(Paragraph("3. История отсутствует", heading_style))
                story.append(Spacer(1, 0.5*cm))
            
            # 4. Настройки уведомлений
            story.append(Paragraph("4. Настройки уведомлений", heading_style))
            if notification_settings:
                notif_data = [
                    ['Параметр', 'Значение'],
                    ['Ежедневное расписание', 'Да' if notification_settings.get('daily_schedule') else 'Нет'],
                    ['Время отправки', notification_settings.get('daily_schedule_time', 'Не настроено')],
                    ['Напоминания о парах', 'Да' if notification_settings.get('lesson_reminder') else 'Нет'],
                    ['За сколько минут', str(notification_settings.get('lesson_reminder_minutes', '-'))],
                    ['Уведомления о пропусках', 'Да' if notification_settings.get('hours_notification') else 'Нет'],
                    ['Порог пропусков', str(notification_settings.get('hours_threshold', '-'))],
                ]
                
                notif_table = Table(notif_data, colWidths=[8*cm, 10*cm])
                notif_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(notif_table)
            else:
                story.append(Paragraph("Настройки не найдены", normal_style))
            
            story.append(Spacer(1, 0.5*cm))
            
            # 5. Статистика
            story.append(Paragraph("5. Статистика", heading_style))
            stats_data = [
                ['Параметр', 'Значение'],
                ['Всего записей', str(len(hours_history))],
                ['Максимум', str(max([r[1] for r in hours_history], default=0))],
                ['Минимум', str(min([r[1] for r in hours_history], default=0))],
            ]
            
            stats_table = Table(stats_data, colWidths=[8*cm, 10*cm])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(stats_table)
            
            # Футер
            story.append(Spacer(1, 1*cm))
            story.append(Paragraph("--- MTC Helper Bot ---", normal_style))
            
            # Сохранение PDF
            doc.build(story)
            logger.info(f"PDF файл успешно создан: {filepath}")
            
            return filepath
        
        except Exception as e:
            logger.error(f"Ошибка экспорта всех данных пользователя: {e}", exc_info=True)
            raise


# Глобальный экземпляр
excel_exporter = ExcelExporter()
