"""
Система управления пользователями для администратора
"""
import asyncio
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any, Tuple
from aiogram import types
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from utils.db import DB
from utils.log import create_logger
from utils.admin_logger import get_admin_logger
from config import bot

logger = create_logger(__name__)


class UserManager:
    """Менеджер пользователей"""
    
    def __init__(self):
        self.db = DB()
        self.admin_logger = get_admin_logger()
    
    def search_users(self, query: str, search_type: str = 'auto') -> List[Dict[str, Any]]:
        """
        Поиск пользователей
        
        Args:
            query: Поисковый запрос
            search_type: Тип поиска (auto, id, username, group)
        
        Returns:
            Список найденных пользователей
        """
        try:
            users = []
            
            if search_type == 'auto':
                # Автоопределение типа поиска
                if query.startswith('@'):
                    search_type = 'username'
                    query = query[1:]  # Убираем @
                elif query.isdigit():
                    # Если число длинное (>6 цифр), скорее всего это user_id
                    # Если короткое (<=6 цифр), скорее всего это группа
                    if len(query) > 6:
                        search_type = 'id'
                    else:
                        search_type = 'group'
                else:
                    search_type = 'group'
            
            if search_type == 'id':
                # Поиск по user_id
                self.db.cursor.execute(
                    '''SELECT user_id, tg_username, group_id, sec_group_id, missed_hours, 
                              created_at, last_activity, is_blocked
                       FROM users WHERE user_id = ?''',
                    (int(query),)
                )
                rows = self.db.cursor.fetchall()
                
            elif search_type == 'username':
                # Поиск по username
                self.db.cursor.execute(
                    '''SELECT user_id, tg_username, group_id, sec_group_id, missed_hours,
                              created_at, last_activity, is_blocked
                       FROM users WHERE tg_username LIKE ?''',
                    (f'%{query}%',)
                )
                rows = self.db.cursor.fetchall()
                
            elif search_type == 'group':
                # Поиск по группе
                self.db.cursor.execute(
                    '''SELECT user_id, tg_username, group_id, sec_group_id, missed_hours,
                              created_at, last_activity, is_blocked
                       FROM users WHERE group_id = ? OR sec_group_id = ?''',
                    (query, query)
                )
                rows = self.db.cursor.fetchall()
            else:
                return []
            
            for row in rows:
                user = {
                    'user_id': row[0],
                    'username': row[1],
                    'group_id': row[2],
                    'sec_group_id': row[3],
                    'missed_hours': row[4],
                    'created_at': row[5],
                    'last_activity': row[6],
                    'is_blocked': row[7]
                }
                users.append(user)
            
            return users
            
        except Exception as e:
            logger.error(f"Failed to search users: {e}")
            return []
    
    def get_user_profile(self, user_id: int) -> Optional[Dict[str, Any]]:
        """
        Получить полный профиль пользователя
        
        Args:
            user_id: ID пользователя
        
        Returns:
            Словарь с данными пользователя
        """
        try:
            # Основные данные
            self.db.cursor.execute(
                '''SELECT user_id, tg_username, group_id, sec_group_id, missed_hours,
                          show_missed_hours_mode, smena, created_at, last_activity, is_blocked
                   FROM users WHERE user_id = ?''',
                (user_id,)
            )
            row = self.db.cursor.fetchone()
            
            if not row:
                return None
            
            profile = {
                'user_id': row[0],
                'username': row[1],
                'group_id': row[2],
                'sec_group_id': row[3],
                'missed_hours': row[4],
                'show_missed_hours_mode': row[5],
                'smena': row[6],
                'created_at': row[7],
                'last_activity': row[8],
                'is_blocked': row[9]
            }
            
            # Настройки уведомлений
            self.db.cursor.execute(
                '''SELECT daily_schedule, daily_schedule_time, lesson_reminder, 
                          lesson_reminder_minutes, hours_threshold, hours_notification
                   FROM notification_settings WHERE user_id = ?''',
                (user_id,)
            )
            notif_row = self.db.cursor.fetchone()
            
            if notif_row:
                profile['notifications'] = {
                    'daily_schedule': bool(notif_row[0]),
                    'daily_schedule_time': notif_row[1],
                    'lesson_reminder': bool(notif_row[2]),
                    'lesson_reminder_minutes': notif_row[3],
                    'hours_threshold': notif_row[4],
                    'hours_notification': bool(notif_row[5])
                }
            else:
                profile['notifications'] = None
            
            # Проверяем блокировку
            if profile['is_blocked']:
                self.db.cursor.execute(
                    'SELECT blocked_by, reason, blocked_at FROM blocked_users WHERE user_id = ?',
                    (user_id,)
                )
                block_row = self.db.cursor.fetchone()
                if block_row:
                    profile['block_info'] = {
                        'blocked_by': block_row[0],
                        'reason': block_row[1],
                        'blocked_at': block_row[2]
                    }
            
            return profile
            
        except Exception as e:
            logger.error(f"Failed to get user profile: {e}")
            return None
    
    def get_user_activity_history(self, user_id: int, limit: int = 20) -> List[Dict[str, Any]]:
        """
        Получить историю активности пользователя
        
        Args:
            user_id: ID пользователя
            limit: Максимальное количество записей
        
        Returns:
            Список записей активности
        """
        try:
            self.db.cursor.execute(
                '''SELECT action, details, timestamp
                   FROM user_activity
                   WHERE user_id = ?
                   ORDER BY timestamp DESC
                   LIMIT ?''',
                (user_id, limit)
            )
            
            activities = []
            for row in self.db.cursor.fetchall():
                activity = {
                    'action': row[0],
                    'details': row[1],
                    'timestamp': row[2]
                }
                activities.append(activity)
            
            return activities
            
        except Exception as e:
            logger.error(f"Failed to get user activity: {e}")
            return []
    
    def get_user_hours_history(self, user_id: int, limit: int = 20) -> List[Dict[str, Any]]:
        """
        Получить историю изменений пропущенных часов
        
        Args:
            user_id: ID пользователя
            limit: Максимальное количество записей
        
        Returns:
            Список изменений
        """
        try:
            self.db.cursor.execute(
                '''SELECT hours, date
                   FROM hours_history
                   WHERE user_id = ?
                   ORDER BY date DESC
                   LIMIT ?''',
                (user_id, limit)
            )
            
            history = []
            for row in self.db.cursor.fetchall():
                entry = {
                    'hours': row[0],
                    'date': row[1]
                }
                history.append(entry)
            
            return history
            
        except Exception as e:
            logger.error(f"Failed to get hours history: {e}")
            return []
    
    async def send_personal_message(self, user_id: int, message: str, admin_id: int) -> bool:
        """
        Отправить личное сообщение пользователю
        
        Args:
            user_id: ID пользователя
            message: Текст сообщения
            admin_id: ID администратора
        
        Returns:
            True если успешно
        """
        try:
            await bot.send_message(
                chat_id=user_id,
                text=f"📨 <b>Сообщение от администратора:</b>\n\n{message}",
                parse_mode="HTML"
            )
            
            # Логируем действие
            self.admin_logger.log_action(
                admin_id=admin_id,
                action='user_message',
                target_id=user_id,
                details={'message_preview': message[:100]}
            )
            
            logger.info(f"Personal message sent to user {user_id} by admin {admin_id}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to send personal message: {e}")
            return False
    
    def change_user_group(self, user_id: int, group_id: str, is_secondary: bool, admin_id: int) -> bool:
        """
        Изменить группу пользователя
        
        Args:
            user_id: ID пользователя
            group_id: Новая группа
            is_secondary: Это дополнительная группа?
            admin_id: ID администратора
        
        Returns:
            True если успешно
        """
        try:
            column = 'sec_group_id' if is_secondary else 'group_id'
            
            self.db.cursor.execute(
                f'UPDATE users SET {column} = ? WHERE user_id = ?',
                (group_id, user_id)
            )
            self.db.conn.commit()
            
            # Логируем действие
            self.admin_logger.log_action(
                admin_id=admin_id,
                action='user_edit',
                target_id=user_id,
                details={
                    'field': column,
                    'new_value': group_id
                }
            )
            
            logger.info(f"User {user_id} group changed to {group_id} by admin {admin_id}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to change user group: {e}")
            return False
    
    def reset_user_hours(self, user_id: int, admin_id: int) -> bool:
        """
        Сбросить пропущенные часы пользователя
        
        Args:
            user_id: ID пользователя
            admin_id: ID администратора
        
        Returns:
            True если успешно
        """
        try:
            # Получаем текущее значение для логирования
            old_hours = self.db.get_column(user_id, 'missed_hours', 'users')
            
            # Преобразуем в int, обрабатываем None и строки
            try:
                if old_hours is None:
                    old_hours_int = 0
                elif isinstance(old_hours, str):
                    old_hours_int = int(old_hours) if old_hours.isdigit() else 0
                else:
                    old_hours_int = int(old_hours)
            except (ValueError, TypeError):
                old_hours_int = 0
            
            self.db.cursor.execute(
                'UPDATE users SET missed_hours = 0 WHERE user_id = ?',
                (user_id,)
            )
            self.db.conn.commit()
            
            # Добавляем запись в историю
            if old_hours_int > 0:
                self.db.add_hours_history(user_id, -old_hours_int)
            
            # Логируем действие
            self.admin_logger.log_action(
                admin_id=admin_id,
                action='user_edit',
                target_id=user_id,
                details={
                    'field': 'missed_hours',
                    'old_value': old_hours_int,
                    'new_value': 0
                }
            )
            
            logger.info(f"User {user_id} hours reset by admin {admin_id}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to reset user hours: {e}", exc_info=True)
            return False
    
    def block_user(self, user_id: int, admin_id: int, reason: Optional[str] = None) -> bool:
        """
        Заблокировать пользователя
        
        Args:
            user_id: ID пользователя
            admin_id: ID администратора
            reason: Причина блокировки
        
        Returns:
            True если успешно
        """
        try:
            # Обновляем флаг блокировки
            self.db.cursor.execute(
                'UPDATE users SET is_blocked = 1 WHERE user_id = ?',
                (user_id,)
            )
            
            # Добавляем запись в таблицу заблокированных
            self.db.cursor.execute(
                '''INSERT OR REPLACE INTO blocked_users (user_id, blocked_by, reason, blocked_at)
                   VALUES (?, ?, ?, ?)''',
                (user_id, admin_id, reason, datetime.now().isoformat())
            )
            self.db.conn.commit()
            
            # Логируем действие
            self.admin_logger.log_action(
                admin_id=admin_id,
                action='user_block',
                target_id=user_id,
                details={'reason': reason}
            )
            
            logger.info(f"User {user_id} blocked by admin {admin_id}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to block user: {e}")
            return False
    
    def unblock_user(self, user_id: int, admin_id: int) -> bool:
        """
        Разблокировать пользователя
        
        Args:
            user_id: ID пользователя
            admin_id: ID администратора
        
        Returns:
            True если успешно
        """
        try:
            # Обновляем флаг блокировки
            self.db.cursor.execute(
                'UPDATE users SET is_blocked = 0 WHERE user_id = ?',
                (user_id,)
            )
            
            # Удаляем из таблицы заблокированных
            self.db.cursor.execute(
                'DELETE FROM blocked_users WHERE user_id = ?',
                (user_id,)
            )
            self.db.conn.commit()
            
            # Логируем действие
            self.admin_logger.log_action(
                admin_id=admin_id,
                action='user_unblock',
                target_id=user_id
            )
            
            logger.info(f"User {user_id} unblocked by admin {admin_id}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to unblock user: {e}")
            return False
    
    def delete_user(self, user_id: int, admin_id: int) -> bool:
        """
        Удалить пользователя из базы данных
        
        Args:
            user_id: ID пользователя
            admin_id: ID администратора
        
        Returns:
            True если успешно
        """
        try:
            # Удаляем из всех связанных таблиц
            self.db.cursor.execute('DELETE FROM users WHERE user_id = ?', (user_id,))
            self.db.cursor.execute('DELETE FROM notification_settings WHERE user_id = ?', (user_id,))
            self.db.cursor.execute('DELETE FROM hours_history WHERE user_id = ?', (user_id,))
            self.db.cursor.execute('DELETE FROM user_activity WHERE user_id = ?', (user_id,))
            self.db.cursor.execute('DELETE FROM blocked_users WHERE user_id = ?', (user_id,))
            self.db.cursor.execute('DELETE FROM favorites WHERE user_id = ?', (user_id,))
            self.db.conn.commit()
            
            # Логируем действие
            self.admin_logger.log_action(
                admin_id=admin_id,
                action='user_delete',
                target_id=user_id
            )
            
            logger.info(f"User {user_id} deleted by admin {admin_id}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to delete user: {e}")
            return False
    
    def cleanup_inactive_users(self, days: int, admin_id: int) -> int:
        """
        Удалить неактивных пользователей
        
        Args:
            days: Количество дней неактивности
            admin_id: ID администратора
        
        Returns:
            Количество удаленных пользователей
        """
        try:
            cutoff_date = datetime.now() - timedelta(days=days)
            
            # Находим неактивных пользователей
            self.db.cursor.execute(
                '''SELECT user_id FROM users 
                   WHERE (last_activity IS NULL OR last_activity < ?) 
                   AND is_blocked = 0''',
                (cutoff_date.isoformat(),)
            )
            
            inactive_users = [row[0] for row in self.db.cursor.fetchall()]
            count = len(inactive_users)
            
            # Удаляем каждого
            for user_id in inactive_users:
                self.delete_user(user_id, admin_id)
            
            # Логируем массовое действие
            self.admin_logger.log_action(
                admin_id=admin_id,
                action='users_cleanup',
                details={
                    'days': days,
                    'deleted_count': count
                }
            )
            
            logger.info(f"Cleaned up {count} inactive users (>{days} days) by admin {admin_id}")
            return count
            
        except Exception as e:
            logger.error(f"Failed to cleanup inactive users: {e}")
            return 0
    
    def export_users_to_excel(self, filter_params: Optional[Dict[str, Any]] = None) -> str:
        """
        Экспорт списка пользователей в Excel
        
        Args:
            filter_params: Параметры фильтрации
        
        Returns:
            Путь к файлу
        """
        try:
            # Получаем пользователей
            query = '''SELECT user_id, tg_username, group_id, sec_group_id, missed_hours,
                              created_at, last_activity, is_blocked
                       FROM users'''
            params = []
            
            if filter_params:
                conditions = []
                if 'group' in filter_params:
                    conditions.append('(group_id = ? OR sec_group_id = ?)')
                    params.extend([filter_params['group'], filter_params['group']])
                if 'is_blocked' in filter_params:
                    conditions.append('is_blocked = ?')
                    params.append(1 if filter_params['is_blocked'] else 0)
                
                if conditions:
                    query += ' WHERE ' + ' AND '.join(conditions)
            
            query += ' ORDER BY user_id'
            
            self.db.cursor.execute(query, params)
            users = self.db.cursor.fetchall()
            
            # Создаем Excel файл
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Пользователи"
            
            # Заголовки
            headers = ['User ID', 'Username', 'Группа', 'Доп. группа', 'Пропуски (ч)', 
                      'Регистрация', 'Последняя активность', 'Заблокирован']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Данные
            for row_idx, user in enumerate(users, 2):
                ws.cell(row=row_idx, column=1, value=user[0])
                ws.cell(row=row_idx, column=2, value=user[1] or '-')
                ws.cell(row=row_idx, column=3, value=user[2])
                ws.cell(row=row_idx, column=4, value=user[3] or '-')
                ws.cell(row=row_idx, column=5, value=user[4] or 0)
                ws.cell(row=row_idx, column=6, value=user[5] or '-')
                ws.cell(row=row_idx, column=7, value=user[6] or '-')
                ws.cell(row=row_idx, column=8, value='Да' if user[7] else 'Нет')
            
            # Автоширина колонок
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Сохраняем файл
            filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"data/{filename}"
            wb.save(filepath)
            
            logger.info(f"Users exported to {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Failed to export users: {e}")
            return ""
    
    def format_user_profile(self, profile: Dict[str, Any]) -> str:
        """
        Форматирование профиля пользователя для отображения
        
        Args:
            profile: Данные профиля
        
        Returns:
            Отформатированная строка
        """
        text = "👤 <b>Профиль пользователя</b>\n\n"
        
        text += f"<b>ID:</b> <code>{profile['user_id']}</code>\n"
        text += f"<b>Username:</b> @{profile['username'] or 'не указан'}\n"
        text += f"<b>Группа:</b> {profile['group_id']}\n"
        
        if profile['sec_group_id']:
            text += f"<b>Доп. группа:</b> {profile['sec_group_id']}\n"
        
        text += f"<b>Пропуски:</b> {profile['missed_hours'] or 0}ч\n"
        text += f"<b>Смена:</b> {profile['smena']}\n"
        
        if profile['created_at']:
            created = datetime.fromisoformat(profile['created_at']).strftime('%d.%m.%Y %H:%M')
            text += f"<b>Регистрация:</b> {created}\n"
        
        if profile['last_activity']:
            activity = datetime.fromisoformat(profile['last_activity']).strftime('%d.%m.%Y %H:%M')
            text += f"<b>Последняя активность:</b> {activity}\n"
        else:
            text += f"<b>Последняя активность:</b> нет данных\n"
        
        # Статус
        if profile['is_blocked']:
            text += f"\n<b>Статус:</b> 🚫 Заблокирован\n"
            if 'block_info' in profile:
                block_info = profile['block_info']
                blocked_date = datetime.fromisoformat(block_info['blocked_at']).strftime('%d.%m.%Y %H:%M')
                text += f"<b>Дата блокировки:</b> {blocked_date}\n"
                if block_info['reason']:
                    text += f"<b>Причина:</b> {block_info['reason']}\n"
        else:
            text += f"\n<b>Статус:</b> ✅ Активен\n"
        
        # Уведомления
        if profile['notifications']:
            notif = profile['notifications']
            text += f"\n<b>Уведомления:</b>\n"
            text += f"• Ежедневное расписание: {'✅' if notif['daily_schedule'] else '❌'}\n"
            if notif['daily_schedule']:
                text += f"  Время: {notif['daily_schedule_time']}\n"
            text += f"• Напоминания о парах: {'✅' if notif['lesson_reminder'] else '❌'}\n"
            if notif['lesson_reminder']:
                text += f"  За {notif['lesson_reminder_minutes']} мин\n"
            text += f"• Уведомления о пропусках: {'✅' if notif['hours_notification'] else '❌'}\n"
            if notif['hours_notification']:
                text += f"  Порог: {notif['hours_threshold']}ч\n"
        
        return text


# Глобальный экземпляр
_user_manager = None


def get_user_manager() -> UserManager:
    """Получить глобальный экземпляр UserManager"""
    global _user_manager
    if _user_manager is None:
        _user_manager = UserManager()
    return _user_manager
